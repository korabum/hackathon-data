#!flask/bin/python

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Color, Border
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side
from datetime import date
from datetime import datetime, timedelta
import json

border_all_white = Border(top = Side(border_style='thin', color='FFFFFFFF'),    
                right = Side(border_style='thin', color='FFFFFFFF'), 
                bottom = Side(border_style='thin', color='FFFFFFFF'),
                left = Side(border_style='thin', color='FFFFFFFF'))

border_all_black = Border(top = Side(border_style='thin', color='00000000'),    
                right = Side(border_style='thin', color='00000000'), 
                bottom = Side(border_style='thin', color='00000000'),
                left = Side(border_style='thin', color='00000000'))

def set_border(ws, cell_range, border=Border()):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = ws[cell_range]

    try:
    	first_col = int(cell_range.split(":")[0][-2:])-1
    except ValueError:
    	first_col = int(cell_range.split(":")[0][-1:])-1

    try:
    	last_col = int(cell_range.split(":")[1][-2:])
    except ValueError:
    	last_col = int(cell_range.split(":")[1][-1:])

    last_col -= first_col
    first_col = 0

    for i in range(first_col,last_col):
    	for cell in rows[i]:
    		cell.border = border

def get_req():
	request = {'data':'json','h2o_th':'int','log_th':'int'}
	return request

def man_master_summary():
	response = {}
	response['data_type'] = 'json'
	response['field'] = {'data':["id","hit_rules","transaction_id","timestamp","engine_decision","final_decision","merchant_id","h2o_score","bayesia_score","blacklisted","whitelisted","is_fraud","is_fraud_x","is_fraud_y"],'h2o_th':'int','log_th':'int'}
	return json.dumps(response)

def master_summary(request):
	wb = Workbook()
	ranges = ['all']
	dest_filename = 'output/ms_%s.xlsx' % (datetime.strftime(datetime.now(),"%Y%m%d%H%M%S"))

	for i in ranges:
		ws = wb.create_sheet(str(i))

		req_input = 'm1,m2,m3,m4,m5,m6,a1,a2,a3,a4,a5,a6'

		list_deny = ['deny']
		list_otherwise = ['','pending','cancel','partial_refund','expire','refund','chargeback','failure','init','None','None']
		list_settlement = ['accept','settlement','capture','authorize']

		if 'm1' in req_input:
			engine_score_m1 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			final_score_m1 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
		if 'm2' in req_input:
			engine_score_m2 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			final_score_m2 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
		if 'm3' in req_input:
			engine_score_m3 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			final_score_m3 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
		if 'm4' in req_input:
			score_m4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
		if 'm5' in req_input:
			score_m5 = [[0,0],[0,0],[0,0],[0,0],[0,0]]
		if 'm6' in req_input:
			score_m6 = [[0,0],[0,0],[0,0]]
		if 'a1' in req_input:
			score_a1 = [[0,0,0],[0,0,0],[0,0,0]]
		if 'a2' in req_input:
			score_a2 = [[0,0],[0,0],[0,0]]
		if 'a3' in req_input:
			score_a3 = [[0,0],[0,0],[0,0]]
		if 'a4' in req_input:
			h2o_score_a4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			bayesia_score_a4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			combination_score_a4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
		if 'a4' in req_input or 'a5' in req_input:
			whitelist = [0,0,0,0,0,0,0,0,0]
			blacklist = [0,0,0,0,0,0,0,0,0]
		if 'a6' in req_input:
			h2o_score_a6 = [[0,0],[0,0],[0,0]]
			bayesia_score_a6 = [[0,0],[0,0],[0,0]]
			combination_score_a6 = [[0,0],[0,0],[0,0]]

		## ============= LOOP STARTS ====================================== ##

		results = json.loads(request['data'])
		
		if(results != None):
			for row in results[1:]:
				timestamp = datetime.strptime(row[3], '%Y-%m-%d %H:%M:%S')
				final_decision = row[5]
				engine_decision = row[4]
				is_fraud = row[11]
				is_fraud_x = int(row[12])
				is_fraud_y = int(row[13])
				bldate = datetime.strptime('2015-09-07 13:29:09', '%Y-%m-%d %H:%M:%S')
				bl_score = row[8]
				h2o_score = row[7]
				h2o_th = request['h2o_th']
				log_th = request['log_th']

				h2o = 1 if h2o_score > h2o_th else 0
				bayesia = 1 if bl_score > log_th else 0
				combination = 1 if h2o and bayesia else 0

				learning_label = None
				if(is_fraud_x == 0 and is_fraud_y == 0):
					learning_label = 0
				elif is_fraud_x in [1,3,4] or is_fraud_y in [1,2,3,4,6]:
					learning_label = 1
				else:
					learning_label = 2

				findec = None
				if final_decision in list_settlement:
					findec = 0
				elif final_decision in list_otherwise or final_decision == None:
					findec = 1
				elif final_decision in list_deny:
					findec = 2

				engdec = None
				if(engine_decision == 'accept'):
					engdec = 0
				elif(engine_decision == 'challenge'):
					engdec = 1
				elif(engine_decision == 'deny'):
					engdec = 2

				white = None
				black = None
				if(timestamp < bldate):
					if(row[1] != None):
						rules = row[1][5:]
						split_rules = rules.split('-')

						for j in range(0, len(split_rules)):
							split_rules[j] = split_rules[j].strip()

						for rule in split_rules:
							if "CB" in rule:
								black = 1 if engine_decision == 'deny' else 0
								white = 1 if engine_decision == 'accept' else 0

				else:
					black = row[9] and engine_decision == 'deny'
					white = row[10] and engine_decision == 'accept'

				if(white):
					whitelist[engdec + (findec * 3)] += 1
					continue
				elif((final_decision == None or final_decision == 'None') and black):
					blacklist[8] += 1
					continue
				elif(black):
					blacklist[engdec + (findec * 3)] += 1
					continue

				if 'm1' in req_input:
					engine_score_m1[(engdec+(is_fraud_y*3))][combination] += 1
					final_score_m1[(findec+(is_fraud_y*3))][combination] += 1
				if 'm2' in req_input:
					engine_score_m2[(engdec+(is_fraud_x*3))][combination] += 1
					final_score_m2[(findec+(is_fraud_x*3))][combination] += 1
				if 'm3' in req_input:
					engine_score_m3[(engdec+(learning_label*3))][combination] += 1
					final_score_m3[(findec+(learning_label*3))][combination] += 1
				if 'm4' in req_input:
					score_m4[is_fraud_y][combination] += 1
				if 'm5' in req_input:
					score_m5[is_fraud_x][combination] += 1
				if 'm6' in req_input:
					score_m6[learning_label][combination] += 1

				if 'a1' in req_input:
					score_a1[learning_label][engdec] += 1
				if 'a2' in req_input:
					score_a2[engdec][combination] += 1
				if 'a3' in req_input:
					score_a3[findec][combination] += 1
				if 'a4' in req_input:
					h2o_score_a4[engdec + (findec * 3)][h2o] += 1
					bayesia_score_a4[engdec + (findec * 3)][bayesia] += 1
					combination_score_a4[engdec + (findec * 3)][combination] +=1
				if 'a6' in req_input:
					h2o_score_a6[learning_label][h2o] += 1
					bayesia_score_a6[learning_label][bayesia] += 1
					combination_score_a6[learning_label][combination] +=1

		print("Finishing, Writing i: " + str(i) + " to Excel...")

		## ====================== TABEL M1 ========================== ##

		# Set All Border White First	
		set_border(ws, 'A1:CW70', border=border_all_white)

		# Header
		ws['A1'] = 'is_fraud_y'
		ws['B1'] = 'Engine Decision'
		ws['C1'] = 'Machine Learning'
		ws['C2'] = '0'
		ws['D2'] = '1'
		ws['F1'] = 'Final Decision'
		ws['G1'] = 'Machine Learning'
		ws['G2'] = '0'
		ws['H2'] = '1'

		# Sidebar
		ws['A3'] = '0'
		ws['A6'] = '1'
		ws['A9'] = '2'
		ws['A12'] = '3'
		ws['A15'] = '4'
		ws['A18'] = '5'
		ws['A21'] = '6'

		# Sub Sidebar
		ws['B3'] = 'Accept'
		ws['B4'] = 'Challenge'
		ws['B5'] = 'Deny'
		ws['B6'] = 'Accept'
		ws['B7'] = 'Challenge'
		ws['B8'] = 'Deny'
		ws['B9'] = 'Accept'
		ws['B10'] = 'Challenge'
		ws['B11'] = 'Deny'
		ws['B12'] = 'Accept'
		ws['B13'] = 'Challenge'
		ws['B14'] = 'Deny'
		ws['B15'] = 'Accept'
		ws['B16'] = 'Challenge'
		ws['B17'] = 'Deny'
		ws['B18'] = 'Accept'
		ws['B19'] = 'Challenge'
		ws['B20'] = 'Deny'
		ws['B21'] = 'Accept'
		ws['B22'] = 'Challenge'
		ws['B23'] = 'Deny'

		ws['F3'] = 'Settlement'
		ws['F4'] = 'Otherwise'
		ws['F5'] = 'Deny'
		ws['F6'] = 'Settlement'
		ws['F7'] = 'Otherwise'
		ws['F8'] = 'Deny'
		ws['F9'] = 'Settlement'
		ws['F10'] = 'Otherwise'
		ws['F11'] = 'Deny'
		ws['F12'] = 'Settlement'
		ws['F13'] = 'Otherwise'
		ws['F14'] = 'Deny'
		ws['F15'] = 'Settlement'
		ws['F16'] = 'Otherwise'
		ws['F17'] = 'Deny'
		ws['F18'] = 'Settlement'
		ws['F19'] = 'Otherwise'
		ws['F20'] = 'Deny'
		ws['F21'] = 'Settlement'
		ws['F22'] = 'Otherwise'
		ws['F23'] = 'Deny'

		# Merge Cells
		ws.merge_cells('A1:A2')
		ws.merge_cells('B1:B2')
		ws.merge_cells('F1:F2')
		ws.merge_cells('C1:D1')
		ws.merge_cells('G1:H1')
		ws.merge_cells('A3:A5')
		ws.merge_cells('A6:A8')
		ws.merge_cells('A9:A11')
		ws.merge_cells('A12:A14')
		ws.merge_cells('A15:A17')
		ws.merge_cells('A18:A20')
		ws.merge_cells('A21:A23')
		ws.merge_cells('E1:E23')

		set_border(ws, 'A1:H23', border=border_all_black)
		ws['A3'].fill = PatternFill(start_color='FFFF0000',
						end_color='FFFF0000',
						fill_type='solid')

		# Score Values
		for j in range(0,len(engine_score_m1)):
			ws['C'+str(3+j)] = engine_score_m1[j][0]
			ws['D'+str(3+j)] = engine_score_m1[j][1]
			ws['G'+str(3+j)] = final_score_m1[j][0]
			ws['H'+str(3+j)] = final_score_m1[j][1]

			ws['C'+str(3+j)].number_format = "#,###"
			ws['D'+str(3+j)].number_format = "#,###"
			ws['G'+str(3+j)].number_format = "#,###"
			ws['H'+str(3+j)].number_format = "#,###"

		## ====================== TABEL M2 ========================== ##

		# Header
		ws['J1'] = 'is_fraud_x'
		ws['K1'] = 'Engine Decision'
		ws['L1'] = 'Machine Learning'
		ws['L2'] = '0'
		ws['M2'] = '1'
		ws['O1'] = 'Final Decision'
		ws['P1'] = 'Machine Learning'
		ws['P2'] = '0'
		ws['Q2'] = '1'

		# Sidebar
		ws['J3'] = '0'
		ws['J6'] = '1'
		ws['J9'] = '2'
		ws['J12'] = '3'
		ws['J15'] = '4'

		# Sub Sidebar
		ws['K3'] = 'Accept'
		ws['K4'] = 'Challenge'
		ws['K5'] = 'Deny'
		ws['K6'] = 'Accept'
		ws['K7'] = 'Challenge'
		ws['K8'] = 'Deny'
		ws['K9'] = 'Accept'
		ws['K10'] = 'Challenge'
		ws['K11'] = 'Deny'
		ws['K12'] = 'Accept'
		ws['K13'] = 'Challenge'
		ws['K14'] = 'Deny'
		ws['K15'] = 'Accept'
		ws['K16'] = 'Challenge'
		ws['K17'] = 'Deny'

		ws['O3'] = 'Settlement'
		ws['O4'] = 'Otherwise'
		ws['O5'] = 'Deny'
		ws['O6'] = 'Settlement'
		ws['O7'] = 'Otherwise'
		ws['O8'] = 'Deny'
		ws['O9'] = 'Settlement'
		ws['O10'] = 'Otherwise'
		ws['O11'] = 'Deny'
		ws['O12'] = 'Settlement'
		ws['O13'] = 'Otherwise'
		ws['O14'] = 'Deny'
		ws['O15'] = 'Settlement'
		ws['O16'] = 'Otherwise'
		ws['O17'] = 'Deny'

		# Merge Cells
		ws.merge_cells('J1:J2')
		ws.merge_cells('K1:K2')
		ws.merge_cells('L1:M1')
		ws.merge_cells('P1:Q1')
		ws.merge_cells('J3:J5')
		ws.merge_cells('J6:J8')
		ws.merge_cells('J9:J11')
		ws.merge_cells('J12:J14')
		ws.merge_cells('J15:J17')

		set_border(ws, 'J1:Q17', border=border_all_black)

		# Score Values
		for j in range(0,len(engine_score_m2)):
			ws['L'+str(3+j)] = engine_score_m2[j][0]
			ws['M'+str(3+j)] = engine_score_m2[j][1]
			ws['P'+str(3+j)] = final_score_m2[j][0]
			ws['Q'+str(3+j)] = final_score_m2[j][1]

			ws['L'+str(3+j)].number_format = "#,###"
			ws['M'+str(3+j)].number_format = "#,###"
			ws['P'+str(3+j)].number_format = "#,###"
			ws['Q'+str(3+j)].number_format = "#,###"

		## ====================== TABEL M3 ========================== ##

		# Header
		ws['S1'] = 'Learning Label'
		ws['T1'] = 'Engine Decision'
		ws['U1'] = 'Machine Learning'
		ws['U2'] = '0'
		ws['V2'] = '1'
		ws['X1'] = 'Final Decision'
		ws['Y1'] = 'Machine Learning'
		ws['Y2'] = '0'
		ws['Z2'] = '1'

		# Sidebar
		ws['S3'] = '0'
		ws['S6'] = '1'
		ws['S9'] = 'Otherwise'

		# Sub Sidebar
		ws['T3'] = 'Accept'
		ws['T4'] = 'Challenge'
		ws['T5'] = 'Deny'
		ws['T6'] = 'Accept'
		ws['T7'] = 'Challenge'
		ws['T8'] = 'Deny'
		ws['T9'] = 'Accept'
		ws['T10'] = 'Challenge'
		ws['T11'] = 'Deny'

		ws['X3'] = 'Settlement'
		ws['X4'] = 'Otherwise'
		ws['X5'] = 'Deny'
		ws['X6'] = 'Settlement'
		ws['X7'] = 'Otherwise'
		ws['X8'] = 'Deny'
		ws['X9'] = 'Settlement'
		ws['X10'] = 'Otherwise'
		ws['X11'] = 'Deny'

		# Merge Cells
		ws.merge_cells('S1:S2')
		ws.merge_cells('T1:T2')
		ws.merge_cells('U1:V1')
		ws.merge_cells('X1:X2')
		ws.merge_cells('Y1:Z1')

		ws.merge_cells('S3:S5')
		ws.merge_cells('S6:S8')
		ws.merge_cells('S9:S11')

		set_border(ws, 'S1:Z11', border=border_all_black)

		# Score Values
		for j in range(0,len(engine_score_m3)):
			ws['U'+str(3+j)] = engine_score_m3[j][0]
			ws['V'+str(3+j)] = engine_score_m3[j][1]
			ws['Y'+str(3+j)] = final_score_m3[j][0]
			ws['Z'+str(3+j)] = final_score_m3[j][1]

			ws['U'+str(3+j)].number_format = "#,###"
			ws['V'+str(3+j)].number_format = "#,###"
			ws['Y'+str(3+j)].number_format = "#,###"
			ws['Z'+str(3+j)].number_format = "#,###"

		## ====================== TABEL M4 ========================== ##

		# Header
		ws['AB1'] = 'is_fraud_y'
		ws['AC1'] = 'Machine Learning'
		ws['AC2'] = '0'
		ws['AD2'] = '1'

		# Sidebar
		ws['AB3'] = '0'
		ws['AB4'] = '1'
		ws['AB5'] = '2'
		ws['AB6'] = '3'
		ws['AB7'] = '4'
		ws['AB8'] = '5'
		ws['AB9'] = '6'

		# Merge Cells
		ws.merge_cells('AB1:AB2')
		ws.merge_cells('AC1:AD1')

		set_border(ws, 'AB1:AD9', border=border_all_black)

		# Score Values
		for j in range(0,len(score_m4)):
			ws['AC'+str(3+j)] = score_m4[j][0]
			ws['AD'+str(3+j)] = score_m4[j][1]

			ws['AC'+str(3+j)].number_format = "#,###"
			ws['AD'+str(3+j)].number_format = "#,###"

		## ====================== TABEL M5 ========================== ##

		# Header
		ws['AF1'] = 'is_fraud_x'
		ws['AG1'] = 'Machine Learning'
		ws['AG2'] = '0'
		ws['AH2'] = '1'

		# Sidebar
		ws['AF3'] = '0'
		ws['AF4'] = '1'
		ws['AF5'] = '2'
		ws['AF6'] = '3'
		ws['AF7'] = '4'

		# Merge Cells
		ws.merge_cells('AF1:AF2')
		ws.merge_cells('AG1:AH1')

		set_border(ws, 'AF1:AH7', border=border_all_black)

		# Score Values
		for j in range(0,len(score_m5)):
			ws['AG'+str(3+j)] = score_m5[j][0]
			ws['AH'+str(3+j)] = score_m5[j][1]

			ws['AG'+str(3+j)].number_format = "#,###"
			ws['AH'+str(3+j)].number_format = "#,###"

		## ====================== TABEL M6 ========================== ##

		# Header
		ws['AJ1'] = 'Learning Label'
		ws['AK1'] = 'Machine Learning'
		ws['AK2'] = '0'
		ws['AL2'] = '1'

		# Sidebar
		ws['AJ3'] = '0'
		ws['AJ4'] = '1'
		ws['AJ5'] = 'Otherwise'

		# Merge Cells
		ws.merge_cells('AJ1:AJ2')
		ws.merge_cells('AK1:AL1')

		set_border(ws, 'AJ1:AL5', border=border_all_black)

		# Score Values
		for j in range(0,len(score_m6)):
			ws['AK'+str(3+j)] = score_m6[j][0]
			ws['AL'+str(3+j)] = score_m6[j][1]

			ws['AK'+str(3+j)].number_format = "#,###"
			ws['AL'+str(3+j)].number_format = "#,###"

		## ====================== TABEL A1 ========================== ##

		# Header
		ws['AN1'] = 'Learning Label'
		ws['AO1'] = 'Aegis'
		ws['AO2'] = 'Accept'
		ws['AP2'] = 'Challenge'
		ws['AQ2'] = 'Deny'

		# Sidebar
		ws['AN3'] = '0'
		ws['AN4'] = '1'
		ws['AN5'] = 'Otherwise'

		ws['AN7'] = 'Fraud Accuracy'
		ws['AP7'] = "=(AQ4+BR4)/SUM(AO4:AQ4,BR4)"

		# Merge Cells
		ws.merge_cells('AN1:AN2')
		ws.merge_cells('AO1:AQ1')
		ws.merge_cells('AN7:AO7')
		ws.merge_cells('AP7:AQ7')

		set_border(ws, 'AN1:AQ5', border=border_all_black)
		set_border(ws, 'AN7:AQ7', border=border_all_black)

		# Score Values
		for j in range(0,len(score_a1)):
			ws['AO'+str(3+j)] = score_a1[j][0]
			ws['AP'+str(3+j)] = score_a1[j][1]
			ws['AQ'+str(3+j)] = score_a1[j][2]

			ws['AO'+str(3+j)].number_format = "#,###"
			ws['AP'+str(3+j)].number_format = "#,###"
			ws['AQ'+str(3+j)].number_format = "#,###"

		## ====================== TABEL A2 ========================== ##

		# Header
		ws['AS1'] = 'Engine Decision'
		ws['AT1'] = 'Machine Learning'
		ws['AT2'] = '0'
		ws['AU2'] = '1'

		# Sidebar
		ws['AS3'] = 'Accept'
		ws['AS4'] = 'Challenge'
		ws['AS5'] = 'Deny'

		# Merge Cells
		ws.merge_cells('AS1:AS2')
		ws.merge_cells('AT1:AU1')

		set_border(ws, 'AS1:AU5', border=border_all_black)

		# Score Values
		for j in range(0,len(score_a2)):
			ws['AT'+str(3+j)] = score_a2[j][0]
			ws['AU'+str(3+j)] = score_a2[j][1]

			ws['AT'+str(3+j)].number_format = "#,###"
			ws['AU'+str(3+j)].number_format = "#,###"

		## ====================== TABEL A3 ========================== ##

		# Header
		ws['AW1'] = 'Final Decision'
		ws['AX1'] = 'Machine Learning'
		ws['AX2'] = '0'
		ws['AY2'] = '1'

		# Sidebar
		ws['AW3'] = 'Settlement'
		ws['AW4'] = 'Otherwise'
		ws['AW5'] = 'Deny'

		# Merge Cells
		ws.merge_cells('AW1:AW2')
		ws.merge_cells('AX1:AY1')

		set_border(ws, 'AW1:AY5', border=border_all_black)

		# Score Values
		for j in range(0,len(score_a3)):
			ws['AX'+str(3+j)] = score_a3[j][0]
			ws['AY'+str(3+j)] = score_a3[j][1]

			ws['AX'+str(3+j)].number_format = "#,###"
			ws['AY'+str(3+j)].number_format = "#,###"

		## ====================== TABEL A4 ========================== ##

		# Judul Atas
		ws['BA1'] = 'Final Decision'
		ws['BB1'] = 'Engine Decision'
		ws['BC1'] = 'Whitelist'
		ws['BD1'] = 'Blacklist'
		ws['BE1'] = 'H2O'
		ws['BG1'] = 'BayesiaLab'
		ws['BI1'] = 'Combination'

		# Sub Judul Atas
		ws['BE2'] = '0'
		ws['BF2'] = '1'
		ws['BG2'] = '0'
		ws['BH2'] = '1'
		ws['BI2'] = '0'
		ws['BJ2'] = '1'

		# Judul Samping
		ws['BA3'] = 'Settlement'
		ws['BA6'] = 'Otherwise'
		ws['BA9'] = 'Deny'

		# Sub Judul Samping
		ws['BB3'] = 'Accept'
		ws['BB4'] = 'Challenge'
		ws['BB5'] = 'Deny'
		ws['BB6'] = 'Accept'
		ws['BB7'] = 'Challenge'
		ws['BB8'] = 'Deny'
		ws['BB9'] = 'Accept'
		ws['BB10'] = 'Challenge'
		ws['BB11'] = 'Deny'

		# Acceptance Rate
		ws['BA13'] = 'Total Data = 	'
		ws['BA14'] = 'Initial Acceptance Rate = '
		ws['BA15'] = 'Aegis Acceptance rate = '
		ws['BA16'] = 'Final Acceptance Rate = '
		ws['BC16'] = 'Optimis'
		ws['BC17'] = 'Pesimis'
		ws['BA18'] = 'Aegis Final Acceptance Rate = '

		ws['BC13'] = "=SUM(BC3:BC11)"
		ws['BD13'] = "=SUM(BD3:BD11)"
		ws['BE13'] = "=SUM(BE3:BE11)"
		ws['BF13'] = "=SUM(BF3:BF11)"
		ws['BG13'] = "=SUM(BG3:BG11)"
		ws['BH13'] = "=SUM(BH3:BH11)"
		ws['BI13'] = "=SUM(BI3:BI11)"
		ws['BJ13'] = "=SUM(BJ3:BJ11)"

		ws['BE14'] = "=SUM(BC3:BC11,BE3:BE11)/SUM(BC3:BD11,BE3:BF11)"
		ws['BG14'] = "=SUM(BC3:BC11,BG3:BG11)/SUM(BC3:BD11,BG3:BH11)"
		ws['BI14'] = "=SUM(BC3:BC11,BI3:BI11)/SUM(BC3:BD11,BI3:BJ11)"
		ws['BE15'] = "=(SUM(BC3:BC11,BE3:BF3,BE6:BF6,BE9:BF9)+0.5*(SUM(BE4:BF4,BE7:BF7,BE10:BF10)))/SUM(BC3:BF11)"
		ws['BE16'] = "=SUM(BC3:BC5,BE3:BE5,BE7:BE8,BE10:BE11)/SUM(BC3:BD11,BE3:BF11)"
		ws['BG16'] = "=SUM(BC3:BC5,BG3:BG5,BG7:BG8,BG10:BG11)/SUM(BC3:BD11,BG3:BH11)"
		ws['BI16'] = "=SUM(BC3:BC5,BI3:BI5,BI7:BI8,BI10:BI11)/SUM(BC3:BD11,BI3:BJ11)"
		ws['BE17'] = "=SUM(BC3:BC5,BE3:BE5)/SUM(BC3:BD11,BE3:BF11)"
		ws['BG17'] = "=SUM(BC3:BC5,BG3:BG5)/SUM(BC3:BD11,BG3:BH11)"
		ws['BI17'] = "=SUM(BC3:BC5,BI3:BI5)/SUM(BC3:BD11,BI3:BJ11)"
		ws['BE18'] = "=SUM(BC3:BC5,BE3:BF5)/SUM(BC3:BF11)"

		# Merge Cells
		ws.merge_cells('BA1:BA2')
		ws.merge_cells('BB1:BB2')
		ws.merge_cells('BC1:BC2')
		ws.merge_cells('BD1:BD2')
		ws.merge_cells('BE1:BF1')
		ws.merge_cells('BG1:BH1')
		ws.merge_cells('BI1:BJ1')
		ws.merge_cells('BA3:BA5')
		ws.merge_cells('BA6:BA8')
		ws.merge_cells('BA9:BA11')

		ws.merge_cells('BA13:BB13')
		ws.merge_cells('BA14:BD14')
		ws.merge_cells('BA15:BD15')
		ws.merge_cells('BA16:BB17')
		ws.merge_cells('BC16:BD16')
		ws.merge_cells('BC17:BC17')
		ws.merge_cells('BA18:BD18')

		ws.merge_cells('BE14:BF14')
		ws.merge_cells('BG14:BH14')
		ws.merge_cells('BI14:BJ14')
		ws.merge_cells('BE15:BJ15')
		ws.merge_cells('BE16:BF16')
		ws.merge_cells('BG16:BH16')
		ws.merge_cells('BI16:BJ16')
		ws.merge_cells('BE17:BF17')
		ws.merge_cells('BG17:BH17')
		ws.merge_cells('BI17:BJ17')
		ws.merge_cells('BE18:BJ18')

		set_border(ws, 'BA1:BJ11', border=border_all_black)
		set_border(ws, 'BA13:BJ18', border=border_all_black)

		# Score Values
		for j in range(0,len(whitelist)):
			ws['BC'+str(3+j)] = whitelist[j]
			ws['BD'+str(3+j)] = blacklist[j]
			ws['BE'+str(3+j)] = h2o_score_a4[j][0]
			ws['BF'+str(3+j)] = h2o_score_a4[j][1]
			ws['BG'+str(3+j)] = bayesia_score_a4[j][0]
			ws['BH'+str(3+j)] = bayesia_score_a4[j][1]
			ws['BI'+str(3+j)] = combination_score_a4[j][0]
			ws['BJ'+str(3+j)] = combination_score_a4[j][1]

			ws['BC'+str(3+j)].number_format = "#,###"
			ws['BD'+str(3+j)].number_format = "#,###"
			ws['BE'+str(3+j)].number_format = "#,###"
			ws['BF'+str(3+j)].number_format = "#,###"
			ws['BG'+str(3+j)].number_format = "#,###"
			ws['BH'+str(3+j)].number_format = "#,###"
			ws['BI'+str(3+j)].number_format = "#,###"
			ws['BJ'+str(3+j)].number_format = "#,###"

		## ====================== TABEL A5 ========================== ##

		# Judul Atas
		ws['BL1'] = 'Final Decision'
		ws['BM1'] = 'Engine Decision'
		ws['BN1'] = 'Whitelist'
		ws['BO1'] = 'Blacklist'

		# Judul Samping
		ws['BL3'] = 'Settlement'
		ws['BL6'] = 'Otherwise'
		ws['BL9'] = 'Deny'

		# Sub Judul Samping
		ws['BM3'] = 'Accept'
		ws['BM4'] = 'Challenge'
		ws['BM5'] = 'Deny'
		ws['BM6'] = 'Accept'
		ws['BM7'] = 'Challenge'
		ws['BM8'] = 'Deny'
		ws['BM9'] = 'Accept'
		ws['BM10'] = 'Challenge'
		ws['BM11'] = 'Deny'

		# Merge Cells
		ws.merge_cells('BL1:BL2')
		ws.merge_cells('BM1:BM2')
		ws.merge_cells('BM1:BM2')
		ws.merge_cells('BO1:BO2')
		ws.merge_cells('BL3:BL5')
		ws.merge_cells('BL6:BL8')
		ws.merge_cells('BL9:BL11')

		set_border(ws, 'BL1:BO11', border=border_all_black)

		# Score Values
		for j in range(0,len(whitelist)):
			ws['BN'+str(3+j)] = whitelist[j]
			ws['BO'+str(3+j)] = blacklist[j]

			ws['BN'+str(3+j)].number_format = "#,###"
			ws['BO'+str(3+j)].number_format = "#,###"

		## ====================== TABEL A6 ========================== ##

		# Judul Atas
		ws['BQ1'] = 'Learning Label'
		ws['BR1'] = 'Whitelist'
		ws['BR2'] = 'Blacklist'
		ws['BS1'] = 'H2O'
		ws['BU1'] = 'BayesiaLab'
		ws['BW1'] = 'Combination'

		# Sub Judul Atas
		ws['BS2'] = '0'
		ws['BT2'] = '1'
		ws['BU2'] = '0'
		ws['BV2'] = '1'
		ws['BW2'] = '0'
		ws['BX2'] = '1'

		# Judul Samping
		ws['BQ3'] = '0'
		ws['BQ4'] = '1'
		ws['BQ5'] = 'Otherwise'

		# Acceptance Rate
		ws['BR3'] = "=SUM(BC3:BC11)"
		ws['BR4'] = "=SUM(BD3:BD11)"
		ws['BR5'] = '0'

		ws['BQ7'] = 'Fraud Accuracy = 	'
		ws['BQ8'] = 'Acceptance Rate = 	'

		ws['BS7'] = "=SUM(BR4,BT4)/SUM(BR4:BT4)"
		ws['BS8'] = "=SUM(BR3,BS3:BS5)/SUM(BR3:BT5)"
		ws['BU7'] = "=SUM(BR4,BV4)/SUM(BR4,BU4:BV4)"
		ws['BU8'] = "=SUM(BR3,BU3:BU5)/SUM(BR3:BR5,BU3:BV5)"
		ws['BW7'] = "=SUM(BR4,BX4)/SUM(BR4,BW4:BX4)"
		ws['BW8'] = "=SUM(BR3,BW3:BW5)/SUM(BR3:BR5,BW3:BX5)"

		# Merge Cells
		ws.merge_cells('BQ1:BQ2')
		ws.merge_cells('BS1:BT1')
		ws.merge_cells('BU1:BV1')
		ws.merge_cells('BW1:BX1')

		ws.merge_cells('BS7:BT7')
		ws.merge_cells('BU7:BV7')
		ws.merge_cells('BW7:BX7')
		ws.merge_cells('BS8:BT8')
		ws.merge_cells('BU8:BV8')
		ws.merge_cells('BW8:BX8')

		set_border(ws, 'BQ1:BX5', border=border_all_black)
		set_border(ws, 'BQ7:BX8', border=border_all_black)

		# Score Values
		for j in range(0,len(h2o_score_a6)):
			ws['BS'+str(3+j)] = h2o_score_a6[j][0]
			ws['BT'+str(3+j)] = h2o_score_a6[j][1]
			ws['BU'+str(3+j)] = bayesia_score_a6[j][0]
			ws['BV'+str(3+j)] = bayesia_score_a6[j][1]
			ws['BW'+str(3+j)] = combination_score_a6[j][0]
			ws['BX'+str(3+j)] = combination_score_a6[j][1]

			ws['BS'+str(3+j)].number_format = "#,###"
			ws['BT'+str(3+j)].number_format = "#,###"
			ws['BU'+str(3+j)].number_format = "#,###"
			ws['BV'+str(3+j)].number_format = "#,###"
			ws['BW'+str(3+j)].number_format = "#,###"
			ws['BX'+str(3+j)].number_format = "#,###"

		# ====== Label ======
		ws['A25'] = 'M1'
		ws.merge_cells('A25:H25')
		ws['J19'] = 'M2'
		ws.merge_cells('J19:Q19')
		ws['S13'] = 'M3'
		ws.merge_cells('S13:Z13')
		ws['AB11'] = 'M4'
		ws.merge_cells('AB11:AD11')
		ws['AF9'] = 'M5'
		ws.merge_cells('AF9:AH9')
		ws['AJ7'] = 'M6'
		ws.merge_cells('AJ7:AL7')

		ws['AN9'] = 'A1'
		ws.merge_cells('AN9:AQ9')
		ws['AS7'] = 'A2'
		ws.merge_cells('AS7:AU7')
		ws['AW7'] = 'A3'
		ws.merge_cells('AW7:AY7')
		ws['BA20'] = 'A4'
		ws.merge_cells('BA20:BJ20')
		ws['BL13'] = 'A5'
		ws.merge_cells('BL13:BO13')
		ws['BQ10'] = 'A6'
		ws.merge_cells('BQ10:BX10')

		wb.save(filename = dest_filename)

	ws = wb.get_sheet_by_name('Sheet')
	wb.remove_sheet(ws)
	wb.save(filename = dest_filename)

	response = {'status':200, 'output':dest_filename}
	return response
