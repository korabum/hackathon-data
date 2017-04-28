#!/usr/bin/python

import time
import csv
import numpy as np
from datetime import date
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Color, Border
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side
import json

## ======================== EXCEL STYLES ======================== ##

border_all_white = Border(top = Side(border_style='thin', color='FFFFFFFF'),    
                right = Side(border_style='thin', color='FFFFFFFF'), 
                bottom = Side(border_style='thin', color='FFFFFFFF'),
                left = Side(border_style='thin', color='FFFFFFFF'))

border_all_black = Border(top = Side(border_style='thin', color='00000000'),    
                right = Side(border_style='thin', color='00000000'), 
                bottom = Side(border_style='thin', color='00000000'),
                left = Side(border_style='thin', color='00000000'))

def set_border(ws, cell_range, border=Border()):
    rows = ws[cell_range]

    try:
    	first_col = int(cell_range.split(":")[0][-2:])-1
    except ValueError:
    	first_col = int(cell_range.split(":")[0][-1:])-1

    try:
    	last_col = int(cell_range.split(":")[1][-2:])
    except ValueError:
    	last_col = int(cell_range.split(":")[1][-1:])

    last_col -= first_col
    first_col = 0

    for i in range(first_col,last_col):
    	for cell in rows[i]:
    		cell.border = border

## ====================== SETUP SECTION ============================= ##

def execute(request):
	wb = Workbook()
	dest_filename = 'output/srs_%s.csv' % (datetime.strftime(datetime.now(),"%Y%m%d%H%M%S"))
	ws1 = wb.create_sheet('grand_summary')

	# Set description of all rules
	rulesdict = {}
	with open('resources/rules_full_table.csv', 'rb') as f:
		reader = csv.reader(f)
		for rule in reader:
			try:
				rulesdict[rule[4]].extend([[rule[6],rule[1],rule[5]]])
			except KeyError:
				rulesdict[rule[4]] = [[rule[6],rule[1],rule[5]]]

	merchant_list = []
	with open('resources/merchant_list.csv', 'rb') as f:
		reader = csv.reader(f)
		for rule in reader:
			merchant_list.append([rule[0],rule[1]])

	mcht_size = len(merchant_list)

	# Set rules to be released
	release0 = []
	release166 = []
	release642 = []
	release808 = []
	release1180 = []

	# with open('resources/rules_list_166.csv', 'rb') as f:
	#     reader = csv.reader(f)
	#     release166 = np.array(list(reader)).ravel()

	# with open('resources/rules_list_642.csv', 'rb') as f:
	#     reader = csv.reader(f)
	#     release642 = np.array(list(reader)).ravel()

	# with open('resources/rules_list_808.csv', 'rb') as f:
	#     reader = csv.reader(f)
	#     release808 = np.array(list(reader)).ravel()

	# with open('resources/rules_list_1180.csv', 'rb') as f:
	#     reader = csv.reader(f)
	#     release1180 = np.array(list(reader)).ravel()

	with open('resources/rules_list_farec.csv', 'rb') as f:
		reader = csv.reader(f)
		release166 = np.array(list(reader)).ravel()

	with open('resources/rules_list_shrec.csv', 'rb') as f:
		reader = csv.reader(f)
		release642 = np.array(list(reader)).ravel()

	with open('resources/rules_list_AYOPOP.csv', 'rb') as f:
		reader = csv.reader(f)
		release808 = np.array(list(reader)).ravel()

	list_deny = ['deny']
	list_otherwise = ['','pending','cancel','partial_refund','expire','refund','chargeback','failure','init','None','NULL']
	list_settlement = ['accept','settlement','capture','authorize']

	ranges = [1,2,3,4,5,6,7,8,9,10,11,12,'all']

	versionrel = ['0','166','642','808']

	# default : 2014-06-26 13:44:11 to 2016-12-21 19:35:11
	time_low = '2014-06-26 13:44:11'
	time_up = '2016-12-21 19:35:11'

	evaluator_dict = { 'ae':'Aegis', 'ml':'ML Only', 'mlr0':'ML+Rule (No Release)', 'mlr1':'ML+Rule (FA Rec)', 'mlr2':'ML+Rule (SH Rec)','mlr3':'ML+Rule (Ayo Rec)'}

	## ====================== LOOP SECTION ============================= ##
	pointer = 2
	counter = 0
	response = []

	for mcht in merchant_list:
		for i in ranges:
			counter_mcht = 0
			ll1_cnt = 0
			isfraud_cnt = 0
			trx_cnt = 0

			ppf1a_cnt = [0,0,0,0,0,0]
			ppf1b_cnt = [0,0,0,0,0,0]
			ppf1c_cnt = [0,0,0,0,0,0]
			ppf2_cnt = [0,0,0,0,0,0]

			is_fraud_x3 = [0,0,0,0,0,0]
			is_fraud_y3 = [0,0,0,0,0,0]

			whitelist = [0,0,0,0,0,0,0,0,0]
			blacklist = [0,0,0,0,0,0,0,0,0]

			score_a1 = [[0,0,0],[0,0,0],[0,0,0]]

			h2o_score_a4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			bayesia_score_a4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]
			combination_score_a4 = [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],[0,0]]

			h2o_score_a6 = [[0,0],[0,0],[0,0]]
			bayesia_score_a6 = [[0,0],[0,0],[0,0]]
			combination_score_a6 = [[0,0],[0,0],[0,0]]

			s_decision_0 = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
			ll_decision_0 = [[0,0,0],[0,0,0],[0,0,0]]
			s_decision_166 = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
			ll_decision_166 = [[0,0,0],[0,0,0],[0,0,0]]
			s_decision_642 = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
			ll_decision_642 = [[0,0,0],[0,0,0],[0,0,0]]
			s_decision_808 = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
			ll_decision_808 = [[0,0,0],[0,0,0],[0,0,0]]
			s_decision_1180 = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
			ll_decision_1180 = [[0,0,0],[0,0,0],[0,0,0]]

			df_res = None
			results = json.loads(request['data'])

			if(len(results) == 0):
				continue

			for trx in results[1:]:
				final_decision = trx[5]
				engine_decision = trx[4]
				timestamp = datetime.strptime(trx[2], '%Y-%m-%d %H:%M:%S')
				whitelist_db = trx[9]
				blacklist_db = trx[8]
				hit_rules = trx[3]
				is_fraud_x = int(trx[10])
				is_fraud_y = int(trx[11])
				final_sherlock_decision = [None,None,None,None,None]
				datenya = datetime.strptime('2015-09-07 13:29:09', '%Y-%m-%d %H:%M:%S')
				trx_cnt += 1
				white = None
				black = None
				h2o_th = request['h2o_th']
				log_th = request['log_th']
				merchant_id = trx[13]

				if(merchant_id != mcht[0]):
					continue

				# print("Sama")

				counter_mcht += 1

				year = '2016'
				if(i != 'all'):
					string_cur = '0%s'% (i) if i < 10 else '%s' % (i)
					string_next = '0%s' % (i+1) if i+1 < 10 else '%s' % (i+1)
					if(string_next == '13'):
						string_next = '01'
						year = '2017'

					ts_cur = datetime.strptime('%s-%s-01 00:00:00' % (year, string_cur), '%Y-%m-%d %H:%M:%S')
					ts_next = datetime.strptime('%s-%s-01 00:00:00' % (year, string_next), '%Y-%m-%d %H:%M:%S')

					if((timestamp < ts_cur and timestamp >= ts_next)):
						continue

				h2o = 1 if trx[6] > h2o_th else 0
				bayesia = 1 if trx[7] > log_th else 0
				combination = 1 if h2o and bayesia else 0

				learning_label = None
				if(is_fraud_x == 0 and is_fraud_y == 0):
					learning_label = 0
				elif is_fraud_x in [1,3,4] or is_fraud_y in [1,2,3,4,6]:
					learning_label = 1
					ll1_cnt += 1
				else:
					learning_label = 2

				findec = None
				if final_decision in list_settlement:
					findec = 0
				elif final_decision in list_otherwise or final_decision == None:
					findec = 1
				elif final_decision in list_deny:
					findec = 2

				engdec = None
				if(engine_decision == 'accept'):
					engdec = 0
				elif(engine_decision == 'challenge'):
					engdec = 1
				elif(engine_decision == 'deny'):
					engdec = 2

				is_fraud = trx[12]
				isfraud_cnt += is_fraud if is_fraud != None else 0

				if(hit_rules != None):
					rules = hit_rules[5:]
					split_rules = rules.split('-')

					# Clean hit rule strings
					for j in range(0, len(split_rules)):
						split_rules[j] = split_rules[j].strip()

					if(timestamp < datenya):
						for rule in split_rules:
							if(rule[:2] == 'CB' and engine_decision == 'deny'):
								black = 1
							elif(rule[:2] == 'CB' and engine_decision == 'accept'):
								white = 1
					else:
						if(blacklist_db == 1 and engine_decision == 'deny'):
							black = 1
						elif(whitelist_db == 1 and engine_decision == 'accept'):
							white = 1
				else:
					white = 1 if whitelist_db == 1 and engine_decision == 'accept' else 0
					black = 1 if blacklist_db == 1 and engine_decision == 'deny' else 0

				if(white == 1):
					whitelist[engdec + (findec * 3)] += 1
					continue
				elif((final_decision == None or final_decision == 'None') and black == 1):
					blacklist[8] += 1
					continue
				elif(black == 1):
					blacklist[engdec + (findec * 3)] += 1
					continue

				if(learning_label == 1 and combination == 0):
					if engdec in [1,2] and findec in [1,2]:
						ppf1a_cnt[1] += 1
						ppf1b_cnt[1] += 1 if is_fraud_y == 3 else 0
						ppf1c_cnt[1] += 1 if is_fraud_y != 3 and is_fraud_x == 3 else 0

				# Fill The Table
				score_a1[learning_label][engdec] += 1

				h2o_score_a4[engdec + (findec * 3)][h2o] += 1
				bayesia_score_a4[engdec + (findec * 3)][bayesia] += 1
				combination_score_a4[engdec + (findec * 3)][combination] +=1

				h2o_score_a6[learning_label][h2o] += 1
				bayesia_score_a6[learning_label][bayesia] += 1
				combination_score_a6[learning_label][combination] +=1

				if(hit_rules != None):
					# Join rules released
					for ver in range(0, len(versionrel)):
						rules = hit_rules[5:]
						split_rules = rules.split('-')

						# Clean hit rule strings
						for j in range(0, len(split_rules)):
							split_rules[j] = split_rules[j].strip()

						release = eval('release' + versionrel[ver])
						rule_inside = [e.strip() for e in split_rules if e.strip() in release]

						# Voting
						if(black != 1 and white != 1):
							for intem in rule_inside:
								split_rules.remove(intem)

							cnt_ac = 0
							cnt_ch = 0
							cnt_de = 0
							cnt_ob = 0
							cnt_all = 0

							ovr_ac = 0
							ovr_ch = 0
							ovr_de = 0

							# Count rule decision
							for item in split_rules:
								nearestrule = None
								nearesttime = 999999999999

								try:
									# Search for the nearest rule decision
									for rule in rulesdict[item]:
										difftime = timestamp - datetime.strptime(rule[0], '%Y-%m-%d %H:%M:%S')
										difftime1 = float(difftime.seconds + (difftime.days * 24 * 3600))
										if(difftime1 < nearesttime and difftime1 > 0):
											nearesttime = difftime1
											nearestrule = rule

									if(nearestrule != None):
										if(nearestrule[1] == 'accept'):
											cnt_ac += 1
											ovr_ac += int(nearestrule[2])
										elif(nearestrule[1] == 'challenge'):
											cnt_ch += 1
											ovr_ch += int(nearestrule[2])
										elif(nearestrule[1] == 'deny'):
											cnt_de += 1
											ovr_de += int(nearestrule[2])
										cnt_ob += 1 if nearestrule[1] == 'observe' else 0
								except KeyError:
									pass

							cnt_all = cnt_ac + cnt_ob + cnt_ch + cnt_de
							override = ovr_ac + ovr_ch + ovr_de

							if(engine_decision == 'accept'):
								if(ovr_ac > 0):
									sherlock_decision = 'accept'
								elif(cnt_ac > 0):
									sherlock_decision = 'accept'
								elif(cnt_de > 0):
									sherlock_decision = 'deny'
								elif(cnt_ch > 0):
									sherlock_decision = 'deny' if combination else 'challenge'
								elif(cnt_ob > 0 or cnt_all == 0):
									sherlock_decision = 'deny' if combination else 'accept'
							elif(engine_decision == 'deny' or engine_decision == 'challenge'):
								if(ovr_de > 0):
									sherlock_decision = 'deny'
								elif(ovr_ch > 0):
									sherlock_decision = 'deny' if combination else 'challenge'
								elif(ovr_ac > 0):
									sherlock_decision = 'accept'
								elif(cnt_de > 0):
									sherlock_decision = 'deny'
								elif(cnt_ch > 0):
									sherlock_decision = 'deny' if combination else 'challenge'
								elif(cnt_ac > 0):
									sherlock_decision = 'accept'
								elif(cnt_ob > 0 or cnt_all == 0):
									sherlock_decision = 'deny' if combination else 'accept'

							final_sherlock_decision[ver] = sherlock_decision

				else:
					for ver in range(0,len(versionrel)):
						final_sherlock_decision[ver] = 'deny' if combination else 'accept'

				for versi in range(0,len(versionrel)):
					s_dec = None
					ver = versionrel[versi]

					if(ver == '0'):
						s_decision = s_decision_0
						ll_decision = ll_decision_0
					elif(ver == '166'):
						s_decision = s_decision_166
						ll_decision = ll_decision_166
					elif(ver == '642'):
						s_decision = s_decision_642
						ll_decision = ll_decision_642
					elif(ver == '808'):
						s_decision = s_decision_808
						ll_decision = ll_decision_808
					# elif(ver == '1180'):
					# 	s_decision = s_decision_1180
					# 	ll_decision = ll_decision_1180

					if(final_sherlock_decision[versi] == 'accept'):
						s_dec = 0
					elif(final_sherlock_decision[versi] == 'challenge'):
						s_dec = 1
					elif(final_sherlock_decision[versi] == 'deny'):
						s_dec = 2

					s_decision[(3 * findec) + engdec][s_dec] += 1
					ll_decision[learning_label][s_dec] += 1

					s_decision_0 = s_decision if ver == '0' else s_decision_0
					s_decision_166 = s_decision if ver == '166' else s_decision_166
					s_decision_642 = s_decision if ver == '642' else s_decision_642
					s_decision_808 = s_decision if ver == '808' else s_decision_808
					# s_decision_1180 = s_decision if ver == '1180' else s_decision_1180

					ll_decision_0 = ll_decision if ver == '0' else ll_decision_0
					ll_decision_166 = ll_decision if ver == '166' else ll_decision_166
					ll_decision_642 = ll_decision if ver == '642' else ll_decision_642
					ll_decision_808 = ll_decision if ver == '808' else ll_decision_808
					# ll_decision_1180 = ll_decision if ver == '1180' else ll_decision_1180
					if(ver == '166'):
						if(learning_label == 1 and s_dec == 0):
							if engdec in [1,2] and findec in [1,2]:
								ppf1a_cnt[3] += 1
								ppf1b_cnt[3] += 1 if is_fraud_y == 3 else 0
								ppf1c_cnt[3] += 1 if is_fraud_y != 3 and is_fraud_x == 3 else 0
						if(learning_label == 1 and engdec == 2 and s_dec == 1):
							ppf2_cnt[3] += 1 if findec == 2 else 0
					if(ver == '642'):
						if(learning_label == 1 and s_dec == 0):
							if engdec in [1,2] and findec in [1,2]:
								ppf1a_cnt[4] += 1
								ppf1b_cnt[4] += 1 if is_fraud_y == 3 else 0
								ppf1c_cnt[4] += 1 if is_fraud_y != 3 and is_fraud_x == 3 else 0
						if(learning_label == 1 and engdec == 2 and s_dec == 1):
							ppf2_cnt[4] += 1 if findec == 2 else 0
					if(ver == '808'):
						if(learning_label == 1 and s_dec == 0):
							if engdec in [1,2] and findec in [1,2]:
								ppf1a_cnt[5] += 1
								ppf1b_cnt[5] += 1 if is_fraud_y == 3 else 0
								ppf1c_cnt[5] += 1 if is_fraud_y != 3 and is_fraud_x == 3 else 0
						if(learning_label == 1 and engdec == 2 and s_dec == 1):
							ppf2_cnt[5] += 1 if findec == 2 else 0

			# print("Blacklist: " + str(sum(blacklist)))
			# print("Whitelist: " + str(sum(whitelist)))
			# print(str(s_decision_642))
			# print(str(ll_decision_642))

			if(counter_mcht == 0):
				continue

			# Aegis
			try:
				ae_fr_acc = float(score_a1[1][2] + sum(blacklist)) / float(sum(score_a1[1]) + sum(blacklist))
			except ZeroDivisionError:
				ae_fr_acc = None

			ae_acc_rate = float(sum(whitelist) + sum(h2o_score_a4[0]) + sum(h2o_score_a4[3]) + sum(h2o_score_a4[6]) + 0.5 * (sum(h2o_score_a4[1]) + sum(h2o_score_a4[4]) + sum(h2o_score_a4[7]))) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in h2o_score_a4))
			ae_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + sum(h2o_score_a4[0]) + sum(h2o_score_a4[1]) + sum(h2o_score_a4[2])) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in h2o_score_a4))

			# ML Only
			try:
				ml_fr_acc = float(sum(blacklist) + combination_score_a6[1][1]) / float(sum(blacklist) + sum(combination_score_a6[1]))
			except ZeroDivisionError:
				ml_fr_acc = None

			ml_acc_rate = float(sum(whitelist) + sum([x[0] for x in combination_score_a4])) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in combination_score_a4))
			ml_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + combination_score_a4[0][0] + combination_score_a4[1][0] + combination_score_a4[2][0] + combination_score_a4[4][0] + combination_score_a4[5][0] + combination_score_a4[7][0] + combination_score_a4[8][0]) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in combination_score_a4))

			# ML + Rule (0)
			try:
				mlr0_fr_acc = float(sum(blacklist) + ll_decision_0[1][2]) / float(sum(blacklist) + sum(ll_decision_0[1]))
			except ZeroDivisionError:
				mlr0_fr_acc = None

			mlr0_acc_rate = float(sum(whitelist) + sum([x[0] for x in s_decision_0]) + 0.5 * sum([x[1] for x in s_decision_0])) / float(sum(whitelist)+sum(blacklist)+sum(sum (x) for x in s_decision_0))
			mlr0_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + s_decision_0[0][0] + s_decision_0[1][0] + s_decision_0[2][0] + s_decision_0[4][0] + s_decision_0[5][0] + s_decision_0[7][0] + s_decision_0[8][0]) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in s_decision_0))

			# ML + Rule (166)
			try:
				mlr1_fr_acc = float(sum(blacklist) + ll_decision_166[1][2]) / float(sum(blacklist) + sum(ll_decision_166[1]))
			except ZeroDivisionError:
				mlr1_fr_acc = None

			mlr1_acc_rate = float(sum(whitelist) + sum([x[0] for x in s_decision_166]) + 0.5 * sum([x[1] for x in s_decision_166])) / float(sum(whitelist)+sum(blacklist)+sum(sum (x) for x in s_decision_166))
			mlr1_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + s_decision_166[0][0] + s_decision_166[1][0] + s_decision_166[2][0] + s_decision_166[4][0] + s_decision_166[5][0] + s_decision_166[7][0] + s_decision_166[8][0]) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in s_decision_166))

			# ML + Rule (642)
			try:
				mlr2_fr_acc = float(sum(blacklist) + ll_decision_642[1][2]) / float(sum(blacklist) + sum(ll_decision_642[1]))
			except ZeroDivisionError:
				mlr2_fr_acc = None

			mlr2_acc_rate = float(sum(whitelist) + sum([x[0] for x in s_decision_642]) + 0.5 * sum([x[1] for x in s_decision_642])) / float(sum(whitelist)+sum(blacklist)+sum(sum (x) for x in s_decision_642))
			mlr2_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + s_decision_642[0][0] + s_decision_642[1][0] + s_decision_642[2][0] + s_decision_642[4][0] + s_decision_642[5][0] + s_decision_642[7][0] + s_decision_642[8][0]) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in s_decision_642))

			# ML + Rule (808)
			try:
				mlr3_fr_acc = float(sum(blacklist) + ll_decision_808[1][2]) / float(sum(blacklist) + sum(ll_decision_808[1]))
			except ZeroDivisionError:
				mlr3_fr_acc = None

			mlr3_acc_rate = float(sum(whitelist) + sum([x[0] for x in s_decision_808]) + 0.5 * sum([x[1] for x in s_decision_808])) / float(sum(whitelist)+sum(blacklist)+sum(sum (x) for x in s_decision_808))
			mlr3_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + s_decision_808[0][0] + s_decision_808[1][0] + s_decision_808[2][0] + s_decision_808[4][0] + s_decision_808[5][0] + s_decision_808[7][0] + s_decision_808[8][0]) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in s_decision_808))

			# # ML + Rule (1180)
			# try:
			# 	mlr4_fr_acc = float(sum(blacklist) + ll_decision_1180[1][2]) / float(sum(blacklist) + sum(ll_decision_1180[1]))
			# except ZeroDivisionError:
			# 	mlr4_fr_acc = None

			# mlr4_acc_rate = float(sum(whitelist) + sum([x[0] for x in s_decision_1180]) + 0.5 * sum([x[1] for x in s_decision_1180])) / float(sum(whitelist)+sum(blacklist)+sum(sum (x) for x in s_decision_1180))
			# mlr4_facc_rate = float(whitelist[0] + whitelist[1] + whitelist[2] + s_decision_1180[0][0] + s_decision_1180[1][0] + s_decision_1180[2][0] + s_decision_1180[4][0] + s_decision_1180[5][0] + s_decision_1180[7][0] + s_decision_1180[8][0]) / float(sum(whitelist) + sum(blacklist) + sum(sum(x) for x in s_decision_1180))


			response.append(['MID','Merchant Name','Evaluator','Month','Fraud Accuracy','Acceptance Rate','Final Acceptance Rate (Opt.)',\
			'Learning Label 1 Count','Transaction Count','Is Fraud?','PF 1a','PF 1b','PF 1c','PF 2'])

			penghitung = 0

			for evaluator in ['ae','ml','mlr0','mlr1','mlr2','mlr3']:
				response.append([mcht[0],mcht[1],evaluator_dict[evaluator],i,eval(evaluator + '_fr_acc'),eval(evaluator + '_acc_rate'),\
				eval(evaluator + '_facc_rate'),ll1_cnt,trx_cnt,isfraud_cnt,ppf1a_cnt[penghitung],ppf1b_cnt[penghitung],ppf1c_cnt[penghitung],
				ppf2_cnt[penghitung]])
				penghitung += 1
				pointer += 1

		counter += 1

	return response



