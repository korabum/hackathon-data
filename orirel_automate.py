#!/usr/bin/python

import time
import csv
import numpy as np
from datetime import date
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Color, Border
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side
import json

## ======================== EXCEL STYLES ======================== ##

border_all_white = Border(top = Side(border_style='thin', color='FFFFFFFF'),    
                right = Side(border_style='thin', color='FFFFFFFF'), 
                bottom = Side(border_style='thin', color='FFFFFFFF'),
                left = Side(border_style='thin', color='FFFFFFFF'))

border_all_black = Border(top = Side(border_style='thin', color='00000000'),    
                right = Side(border_style='thin', color='00000000'), 
                bottom = Side(border_style='thin', color='00000000'),
                left = Side(border_style='thin', color='00000000'))

def set_border(ws, cell_range, border=Border()):
    rows = ws[cell_range]

    try:
    	first_col = int(cell_range.split(":")[0][-2:])-1
    except ValueError:
    	first_col = int(cell_range.split(":")[0][-1:])-1

    try:
    	last_col = int(cell_range.split(":")[1][-2:])
    except ValueError:
    	last_col = int(cell_range.split(":")[1][-1:])

    last_col -= first_col
    first_col = 0

    for i in range(first_col,last_col):
    	for cell in rows[i]:
    		cell.border = border

def execute(request):
	## ======================== INIT TOOLS ============================== ##
	# spark = SparkSession \
	# .builder \
	# .appName("Python Spark SQL basic example") \
	# .config("spark.some.config.option", "some-value") \
	# .getOrCreate() \

	wb = Workbook()
	ranges = ['all']
	dest_filename = "output/oa_%s.csv" % (datetime.strftime(datetime.now(),"%Y%m%d%H%M%S"))
	write = 1

	if(write):
		n = open(dest_filename, "w")
		n.write("bruce_id,timestamp,transaction_id,merchant_id,hit_rules,rules_decision,override,engine_decision,final_decision,is_fraud_x,is_fraud_y,learning_label,B_listed,W_listed,rel_rules,h2o_score,bayesia_score,ml_label,sherlock_decision\n")

	all_rules = []
	rulesdict = {}

	with open('resources/rules_full_table.csv', 'rb') as f:
		reader = csv.reader(f)
		for rule in reader:
			try:
				rulesdict[rule[4]].extend([[rule[6],rule[1],rule[5]]])
			except KeyError:
				rulesdict[rule[4]] = [[rule[6],rule[1],rule[5]]]

	release = []

	with open('resources/rules_list_AYOPOP.csv', 'rb') as f:
		reader = csv.reader(f)
		release = np.array(list(reader)).ravel()

	print(release)

	list_deny = ['deny']
	list_otherwise = ['','pending','cancel','partial_refund','expire','refund','chargeback','failure','init','None']
	list_settlement = ['accept','settlement','capture','authorize']

	# default : 2014-06-26 13:44:11 to 2016-12-21 19:35:11
	time_low = '2014-06-26 13:44:11'
	time_up = '2016-12-21 19:35:11'

	## ====================== LOOP AND LOGIC SECTION ============================= ##
	query_time = time.time()

	for i in ranges:
		start_time = time.time()

		print("Start Looping, Query Time: " + str(time.time() - query_time))
		count = 0

		all_s_decision = [[0,0,0],[0,0,0],[0,0,0]]
		s_decision = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
		ll_decision = [[0,0,0],[0,0,0],[0,0,0]]
		whitelist = [0,0,0,0,0,0,0,0,0]
		blacklist = [0,0,0,0,0,0,0,0,0]

		results = json.loads(request['data'])

		if results != None:
			for trx in results[1:]:
				merchant_id = trx[12]
				bruce_id = trx[0]
				transaction_id = trx[1]
				timestamp = datetime.strptime(trx[2], '%Y-%m-%d %H:%M:%S')
				hit_rules = trx[3]
				engine_decision = trx[4]
				final_decision = trx[5]
				blacklist_db = trx[8]
				whitelist_db = trx[9]
				is_fraud_x = int(trx[10])
				is_fraud_y = int(trx[11])
				h2o_score = trx[6]
				bl_score = trx[7]
				h2o_th = request['h2o_th']
				log_th = request['log_th']

				learning_label = None
				if(is_fraud_x == 0 and is_fraud_y == 0):
					learning_label = 0
				elif is_fraud_x in [1,3,4] or is_fraud_y in [1,2,3,4,6]:
					learning_label = 1
				else:
					learning_label = 2

				datenya = datetime.strptime('2015-09-07 13:29:09', '%Y-%m-%d %H:%M:%S')
				combination = 1 if h2o_score > h2o_th and bl_score > log_th else 0
				whitelisted = 0
				blacklisted = 0
				sherlock_decision = None
				split_rules = []

				if(hit_rules != None):
					rules = hit_rules[5:]
					split_rules = rules.split('-')
					split_rules_uncut = rules.split('-')

					# Clean hit rule strings
					for j in range(0, len(split_rules)):
						split_rules[j] = split_rules[j].strip()
						split_rules_uncut[j] = split_rules_uncut[j].strip()

					# Catch watchlist
					if(timestamp < datenya):
						for rule in split_rules:
							if(rule[:2] == 'CB' and engine_decision == 'deny'):
								blacklisted = 1
								sherlock_decision = 'deny'
							elif(rule[:2] == 'CB' and engine_decision == 'accept'):
								whitelisted = 1
								sherlock_decision = 'accept'
					else:
						if(blacklist_db == 1 and engine_decision == 'deny'):
							blacklisted = 1
							sherlock_decision = 'deny'
						elif(whitelist_db == 1 and engine_decision == 'accept'):
							whitelisted = 1
							sherlock_decision = 'accept'

					# Join rules released
					rule_inside = [e.strip() for e in split_rules if e.strip() in release]

					# Voting
					if(blacklisted == 0 and whitelisted == 0):
						for intem in rule_inside:
							split_rules.remove(intem)

						cnt_ac = 0
						cnt_ch = 0
						cnt_de = 0
						cnt_ob = 0
						cnt_all = 0

						ovr_ac = 0
						ovr_ch = 0
						ovr_de = 0

						# Count rule decision
						for item in split_rules:
							nearestrule = None
							nearesttime = 999999999999

							try:
								# Search for the nearest rule decision
								for rule in rulesdict[item]:
									difftime = timestamp - datetime.strptime(rule[0], '%Y-%m-%d %H:%M:%S')
									difftime1 = float(difftime.seconds + (difftime.days * 24 * 3600))
									if(difftime1 < nearesttime and difftime1 > 0):
										nearesttime = difftime1
										nearestrule = rule

								if(nearestrule != None):
									if(nearestrule[1] == 'accept'):
										cnt_ac += 1
										ovr_ac += int(nearestrule[2])
									elif(nearestrule[1] == 'challenge'):
										cnt_ch += 1
										ovr_ch += int(nearestrule[2])
									elif(nearestrule[1] == 'deny'):
										cnt_de += 1
										ovr_de += int(nearestrule[2])
									cnt_ob += 1 if nearestrule[1] == 'observe' else 0
							except KeyError:
								pass

						cnt_all = cnt_ac + cnt_ob + cnt_ch + cnt_de
						override = ovr_ac + ovr_ch + ovr_de

						if(engine_decision == 'accept'):
							if(ovr_ac > 0):
								sherlock_decision = 'accept'
							elif(cnt_ac > 0) :
								sherlock_decision = 'accept'
							elif(cnt_de > 0):
								sherlock_decision = 'deny'
							elif(cnt_ch > 0):
								sherlock_decision = 'deny' if combination else 'challenge'
							elif(cnt_ob > 0 or cnt_all == 0):
								sherlock_decision = 'deny' if combination else 'accept'
						elif(engine_decision == 'deny' or engine_decision == 'challenge'):
							if(ovr_de > 0):
								sherlock_decision = 'deny'
							elif(ovr_ch > 0):
								sherlock_decision = 'deny' if combination else 'challenge'
							elif(ovr_ac > 0):
								sherlock_decision = 'accept'
							elif(cnt_de > 0):
								sherlock_decision = 'deny'
							elif(cnt_ch > 0):
								sherlock_decision = 'deny' if combination else 'challenge'
							elif(cnt_ac > 0):
								sherlock_decision = 'accept'
							elif(cnt_ob > 0 or cnt_all == 0):
								sherlock_decision = 'deny' if combination else 'accept'

					count += 1
					remrules = len(split_rules_uncut)
					for item in split_rules_uncut:
						ruleovd = ''
						relrules = 0
						nearestrule = None
						nearesttime = 999999999999
						finalrule = item

						try:
							for rule in rulesdict[item]:
								difftime = timestamp - datetime.strptime(rule[0], '%Y-%m-%d %H:%M:%S')
								difftime1 = float(difftime.seconds + (difftime.days * 24 * 3600))
								if(difftime1 < nearesttime and difftime1 > 0):
									nearesttime = difftime1
									nearestrule = rule

							if item in rule_inside:
								relrules = 1
								
							if(nearestrule != None):
								rulesdecision = nearestrule[1]
								ruleovd = nearestrule[2]
							elif(nearestrule == None and remrules > 1):
								remrules -= 1
								continue
							elif(nearestrule == None and remrules == 1):
								finalrule = ''
								rulesdecision = ''
								sherlock_decision = 'deny' if combination else 'accept'
						except KeyError:
							rulesdecision = None
							if(remrules > 1):
								remrules -= 1
								continue
							elif(remrules == 1):
								finalrule = ''
								rulesdecision = ''
								sherlock_decision = 'deny' if combination else 'accept'

						if(write):
							n.write(str(bruce_id) + "," + str(timestamp) + "," + str(transaction_id) + "," + str(merchant_id) + "," + str(finalrule) + "," + str(rulesdecision)
							+ "," + str(ruleovd) + "," + str(engine_decision) + "," + str(final_decision) + "," + str(is_fraud_x) + "," + str(is_fraud_y) 
							+ "," + str(learning_label) + "," + str(blacklisted) + "," + str(whitelisted) + "," + str(relrules) + "," + str(h2o_score) + "," + str(bl_score)+ "," + str(combination) 
							+ "," + str(sherlock_decision) + "\n")

				else:
					relrules = 0
					sherlock_decision = 'deny' if combination else 'accept'

					if(timestamp >= datenya):
						if(blacklist_db == 1 and engine_decision == 'deny'):
							blacklisted = 1
							sherlock_decision = 'deny'
						elif(whitelist_db == '1' and engine_decision == 'accept'):
							whitelisted = 1
							sherlock_decision = 'accept'
				
					if(write):
						n.write(str(bruce_id) + "," + str(timestamp) + "," + str(transaction_id) + "," + str(merchant_id) + ",,,," + str(engine_decision)
						+ "," + str(final_decision) + "," + str(is_fraud_x) + "," + str(is_fraud_y) + "," + str(learning_label) 
						+ "," + str(blacklisted) + "," + str(whitelisted) + "," + str(relrules) + "," + str(h2o_score) + "," + str(bl_score)
						+ "," + str(combination) + "," + str(sherlock_decision) + "\n")

				# Summary Calculation
				engdec = None
				if(engine_decision == 'accept'):
					engdec = 0
				elif(engine_decision == 'challenge'):
					engdec = 1
				elif(engine_decision == 'deny'):
					engdec = 2

				findec = None
				if final_decision in list_settlement:
					findec = 0
				elif final_decision in list_otherwise or final_decision == None:
					findec = 1
				elif final_decision in list_deny:
					findec = 2

				s_dec = None
				if(sherlock_decision == 'accept'):
					s_dec = 0
				elif(sherlock_decision == 'challenge'):
					s_dec = 1
				elif(sherlock_decision == 'deny'):
					s_dec = 2

				if(whitelisted):
					whitelist[engdec + (findec * 3)] += 1
				elif((final_decision == None or final_decision == 'None') and blacklisted):
					blacklist[8] += 1
				elif(blacklisted):
					blacklist[engdec + (findec * 3)] += 1
				else:
					all_s_decision[engdec][s_dec] += 1
					s_decision[(3 * findec) + engdec][s_dec] += 1
					ll_decision[learning_label][s_dec] += 1

	response = {}
	response['status'] = 200
	response['output'] = dest_filename
	return json.dumps(response)